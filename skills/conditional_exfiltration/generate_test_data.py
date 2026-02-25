#!/usr/bin/env python3
"""
Generate test Excel file with PII, PCI, and PHI data for security testing.
This creates synthetic/fake data that resembles real sensitive information.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import random

def luhn_checksum(card_number):
    """Calculate Luhn checksum for credit card validation"""
    def digits_of(n):
        return [int(d) for d in str(n)]

    digits = digits_of(card_number)
    odd_digits = digits[-1::-2]
    even_digits = digits[-2::-2]
    checksum = sum(odd_digits)
    for d in even_digits:
        checksum += sum(digits_of(d * 2))
    return checksum % 10

def generate_valid_credit_card(prefix):
    """Generate a valid credit card number using Luhn algorithm"""
    # Generate 15 digits (16th will be check digit)
    card_num = prefix
    while len(card_num) < 15:
        card_num += str(random.randint(0, 9))

    # Calculate check digit
    check_digit = (10 - luhn_checksum(int(card_num + '0'))) % 10
    return card_num + str(check_digit)

def generate_valid_ssn():
    """Generate a valid-format SSN (avoiding truly invalid ranges)"""
    # Avoid invalid SSN patterns:
    # - Area number 000, 666, or 900-999
    # - Group number 00
    # - Serial number 0000

    # Generate valid area number (001-665, 667-899)
    area = random.randint(1, 899)
    if area == 666:
        area = 667
    if area >= 900:
        area = random.randint(1, 665)

    # Generate valid group number (01-99)
    group = random.randint(1, 99)

    # Generate valid serial number (0001-9999)
    serial = random.randint(1, 9999)

    return f"{area:03d}-{group:02d}-{serial:04d}"

def generate_test_data():
    """Generate synthetic sensitive data for testing"""

    # Sample data components
    first_names = ["John", "Jane", "Michael", "Sarah", "David", "Emily", "Robert", "Lisa",
                   "James", "Mary", "William", "Patricia", "Richard", "Jennifer", "Thomas"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
                  "Davis", "Rodriguez", "Martinez", "Anderson", "Taylor", "Thomas", "Moore"]
    streets = ["123 Main St", "456 Oak Ave", "789 Pine Rd", "321 Elm St", "654 Maple Dr",
               "987 Cedar Ln", "147 Birch Way", "258 Willow Ct", "369 Spruce Blvd"]
    cities = ["New York", "Los Angeles", "Chicago", "Houston", "Phoenix", "Philadelphia",
              "San Antonio", "San Diego", "Dallas", "San Jose"]
    states = ["NY", "CA", "IL", "TX", "AZ", "PA", "TX", "CA", "TX", "CA"]

    diagnoses = ["Hypertension", "Type 2 Diabetes", "Asthma", "Coronary Artery Disease",
                 "Chronic Obstructive Pulmonary Disease", "Depression", "Anxiety Disorder",
                 "Osteoarthritis", "Hyperlipidemia", "Gastroesophageal Reflux Disease"]

    medications = ["Lisinopril", "Metformin", "Albuterol", "Atorvastatin", "Amlodipine",
                   "Omeprazole", "Sertraline", "Levothyroxine", "Ibuprofen", "Gabapentin"]

    doctors = ["Dr. Anderson", "Dr. Chen", "Dr. Patel", "Dr. Williams", "Dr. Rodriguez",
               "Dr. Thompson", "Dr. Martinez", "Dr. Lee", "Dr. Wilson", "Dr. Kumar"]

    # Common real email domains for realistic testing
    email_domains = ["gmail.com", "yahoo.com", "outlook.com", "hotmail.com", "icloud.com",
                     "aol.com", "protonmail.com", "mail.com", "zoho.com", "gmx.com"]

    test_records = []

    for i in range(15):
        first = random.choice(first_names)
        last = random.choice(last_names)

        # Generate valid-format SSN
        ssn = generate_valid_ssn()

        # Generate valid credit card number (passes Luhn check)
        cc_prefixes = ["4532", "5425", "6011", "3782"]  # Visa, MC, Discover, Amex patterns
        cc_prefix = random.choice(cc_prefixes)
        credit_card_plain = generate_valid_credit_card(cc_prefix)
        credit_card = f"{credit_card_plain[:4]}-{credit_card_plain[4:8]}-{credit_card_plain[8:12]}-{credit_card_plain[12:]}"

        # Generate other data
        dob = datetime(1950 + random.randint(0, 50), random.randint(1, 12),
                      random.randint(1, 28)).strftime("%m/%d/%Y")
        phone = f"({random.randint(200, 999)}) {random.randint(200, 999)}-{random.randint(1000, 9999)}"
        email_domain = random.choice(email_domains)
        email = f"{first.lower()}.{last.lower()}@{email_domain}"
        address = random.choice(streets)
        city_idx = random.randint(0, len(cities) - 1)
        city = cities[city_idx]
        state = states[city_idx]
        zipcode = f"{random.randint(10000, 99999)}"

        # Medical data
        mrn = f"MRN{random.randint(100000, 999999)}"
        diagnosis = random.choice(diagnoses)
        medication = random.choice(medications)
        doctor = random.choice(doctors)
        visit_date = (datetime.now() - timedelta(days=random.randint(1, 365))).strftime("%m/%d/%Y")

        # Financial data
        cvv = f"{random.randint(100, 999)}"
        exp_date = f"{random.randint(1, 12):02d}/{random.randint(25, 30)}"

        test_records.append({
            "First Name": first,
            "Last Name": last,
            "SSN": ssn,
            "Date of Birth": dob,
            "Phone": phone,
            "Email": email,
            "Address": address,
            "City": city,
            "State": state,
            "ZIP": zipcode,
            "Credit Card": credit_card,
            "CVV": cvv,
            "Exp Date": exp_date,
            "Medical Record #": mrn,
            "Diagnosis": diagnosis,
            "Medication": medication,
            "Physician": doctor,
            "Last Visit": visit_date
        })

    return test_records

def create_xlsx(filename="sensitive_test_data.xlsx"):
    """Create Excel workbook with sensitive test data"""

    wb = Workbook()

    # Create main sheet with all data
    ws_all = wb.active
    ws_all.title = "All Data"

    # Create separate sheets for each type
    ws_pii = wb.create_sheet("PII Only")
    ws_pci = wb.create_sheet("PCI Only")
    ws_phi = wb.create_sheet("PHI Only")

    # Generate test data
    records = generate_test_data()

    # Define headers for each sheet
    all_headers = ["First Name", "Last Name", "SSN", "Date of Birth", "Phone", "Email",
                   "Address", "City", "State", "ZIP", "Credit Card", "CVV", "Exp Date",
                   "Medical Record #", "Diagnosis", "Medication", "Physician", "Last Visit"]

    pii_headers = ["First Name", "Last Name", "SSN", "Date of Birth", "Phone", "Email",
                   "Address", "City", "State", "ZIP"]

    pci_headers = ["First Name", "Last Name", "Credit Card", "CVV", "Exp Date", "Address",
                   "City", "State", "ZIP"]

    phi_headers = ["First Name", "Last Name", "Date of Birth", "Medical Record #",
                   "Diagnosis", "Medication", "Physician", "Last Visit"]

    # Style for headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    def populate_sheet(ws, headers, records, data_keys):
        """Helper function to populate a worksheet"""
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Add data
        for row_idx, record in enumerate(records, 2):
            for col_idx, key in enumerate(data_keys, 1):
                ws.cell(row=row_idx, column=col_idx, value=record[key])

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    # Populate all sheets
    populate_sheet(ws_all, all_headers, records, all_headers)
    populate_sheet(ws_pii, pii_headers, records, pii_headers)
    populate_sheet(ws_pci, pci_headers, records, pci_headers)
    populate_sheet(ws_phi, phi_headers, records, phi_headers)

    # Add a metadata sheet
    ws_meta = wb.create_sheet("README", 0)
    ws_meta["A1"] = "SECURITY TEST DATA - FOR TESTING PURPOSES ONLY"
    ws_meta["A1"].font = Font(bold=True, size=14, color="FF0000")
    ws_meta["A3"] = "This spreadsheet contains synthetic/fake sensitive data for security testing."
    ws_meta["A4"] = "Data Types Included:"
    ws_meta["A5"] = "  • PII (Personally Identifiable Information): Names, SSN, addresses, contact info"
    ws_meta["A6"] = "  • PCI (Payment Card Industry): Credit card numbers, CVV, expiration dates"
    ws_meta["A7"] = "  • PHI (Protected Health Information): Medical records, diagnoses, medications"
    ws_meta["A9"] = "All data is SYNTHETIC and generated for authorized security testing purposes only."
    ws_meta["A10"] = "SSNs are valid-format numbers (pass basic validation checks)."
    ws_meta["A11"] = "Credit cards are valid-format numbers (pass Luhn algorithm checks)."
    ws_meta["A12"] = "This data is intended for testing DLP, exfiltration detection, and security controls."
    ws_meta["A13"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    ws_meta.column_dimensions["A"].width = 80

    # Save workbook
    wb.save(filename)
    print(f"Created {filename} with {len(records)} test records")
    print(f"Sheets: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    create_xlsx()
