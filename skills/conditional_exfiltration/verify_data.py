#!/usr/bin/env python3
"""
Verify that the generated test data passes validation checks.
"""

from openpyxl import load_workbook
import re

def validate_ssn(ssn):
    """Check if SSN is in valid format and range"""
    # Remove dashes for checking
    parts = ssn.split('-')
    if len(parts) != 3:
        return False, "Invalid format"

    area, group, serial = parts

    # Check format
    if not (len(area) == 3 and len(group) == 2 and len(serial) == 4):
        return False, "Invalid length"

    # Check if all numeric
    if not (area.isdigit() and group.isdigit() and serial.isdigit()):
        return False, "Non-numeric characters"

    area_num = int(area)
    group_num = int(group)
    serial_num = int(serial)

    # Check invalid ranges
    if area_num == 0 or area_num == 666 or area_num >= 900:
        return False, f"Invalid area number: {area_num}"

    if group_num == 0:
        return False, "Invalid group number: 00"

    if serial_num == 0:
        return False, "Invalid serial number: 0000"

    return True, "Valid"

def luhn_check(card_number):
    """Validate credit card using Luhn algorithm"""
    def digits_of(n):
        return [int(d) for d in str(n)]

    # Remove dashes
    card_number = card_number.replace('-', '')

    if not card_number.isdigit():
        return False, "Non-numeric characters"

    if len(card_number) != 16:
        return False, f"Invalid length: {len(card_number)}"

    digits = digits_of(card_number)
    odd_digits = digits[-1::-2]
    even_digits = digits[-2::-2]

    checksum = sum(odd_digits)
    for d in even_digits:
        checksum += sum(digits_of(d * 2))

    is_valid = checksum % 10 == 0
    return is_valid, "Passes Luhn check" if is_valid else "Fails Luhn check"

def verify_xlsx(filename="sensitive_test_data.xlsx"):
    """Verify data in the Excel file"""
    print(f"Verifying {filename}...\n")

    wb = load_workbook(filename)
    ws = wb["All Data"]

    # Find column indices
    headers = [cell.value for cell in ws[1]]
    ssn_col = headers.index("SSN") + 1
    cc_col = headers.index("Credit Card") + 1

    print("=" * 80)
    print("SSN VALIDATION")
    print("=" * 80)

    ssn_valid_count = 0
    ssn_invalid_count = 0

    for row in range(2, ws.max_row + 1):
        ssn = ws.cell(row=row, column=ssn_col).value
        is_valid, message = validate_ssn(ssn)

        if is_valid:
            ssn_valid_count += 1
            print(f"✓ {ssn} - {message}")
        else:
            ssn_invalid_count += 1
            print(f"✗ {ssn} - {message}")

    print(f"\nSSN Summary: {ssn_valid_count} valid, {ssn_invalid_count} invalid")

    print("\n" + "=" * 80)
    print("CREDIT CARD VALIDATION (Luhn Algorithm)")
    print("=" * 80)

    cc_valid_count = 0
    cc_invalid_count = 0

    for row in range(2, ws.max_row + 1):
        cc = ws.cell(row=row, column=cc_col).value
        is_valid, message = luhn_check(cc)

        if is_valid:
            cc_valid_count += 1
            print(f"✓ {cc} - {message}")
        else:
            cc_invalid_count += 1
            print(f"✗ {cc} - {message}")

    print(f"\nCredit Card Summary: {cc_valid_count} valid, {cc_invalid_count} invalid")

    print("\n" + "=" * 80)
    print("OVERALL SUMMARY")
    print("=" * 80)
    print(f"Total records: {ws.max_row - 1}")
    print(f"SSNs passing validation: {ssn_valid_count}/{ws.max_row - 1}")
    print(f"Credit cards passing Luhn check: {cc_valid_count}/{ws.max_row - 1}")

    if ssn_valid_count == ws.max_row - 1 and cc_valid_count == ws.max_row - 1:
        print("\n✓ All data passes validation checks - suitable for security testing!")
    else:
        print("\n✗ Some data failed validation checks")

if __name__ == "__main__":
    verify_xlsx()
